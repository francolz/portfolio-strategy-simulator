"""Utility per esportazione Excel e CSV dalla dashboard Streamlit."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils.helpers import clean_numeric_dataframe


HEADER_FILL = "1F4E78"
HEADER_FONT = "FFFFFF"


def dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    """Converte un DataFrame in CSV UTF-8 scaricabile."""
    return clean_numeric_dataframe(df).to_csv(index=True).encode("utf-8-sig")


def create_excel_export(
    dashboard_summary: pd.DataFrame,
    pac_history: pd.DataFrame,
    buy_hold_history: pd.DataFrame,
    time_series: pd.DataFrame,
    risk_metrics: pd.DataFrame,
    strategy_comparison: pd.DataFrame,
    configuration: pd.DataFrame | None = None,
) -> bytes:
    """Crea un file Excel multi-sheet con tutti i risultati della simulazione.

    Sheet generati:
    1. Dashboard sintetica
    2. Storico PAC
    3. Storico Buy and Hold
    4. Serie temporali
    5. Metriche di rischio
    6. Confronto strategie
    7. Configurazione, se fornita
    """
    output = BytesIO()

    sheets: dict[str, pd.DataFrame] = {
        "Dashboard sintetica": dashboard_summary,
        "Storico PAC": pac_history,
        "Storico BuyHold": buy_hold_history,
        "Serie temporali": time_series,
        "Metriche rischio": risk_metrics,
        "Confronto strategie": strategy_comparison,
    }

    if configuration is not None and not configuration.empty:
        sheets["Configurazione"] = configuration

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            export_df = clean_numeric_dataframe(df.copy())

            # Il foglio Configurazione è già una tabella descrittiva:
            # Sezione | Parametro | Valore.
            # Per questo non esportiamo l'indice.
            write_index = sheet_name != "Configurazione"

            export_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=write_index,
            )

        workbook = writer.book
        for worksheet in workbook.worksheets:
            _style_worksheet(worksheet)

    output.seek(0)
    return output.read()


def _style_worksheet(worksheet) -> None:
    """Applica una formattazione essenziale e leggibile al foglio Excel."""
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_font = Font(color=HEADER_FONT, bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"

            cell.alignment = Alignment(vertical="top", wrap_text=False)

        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            38,
        )